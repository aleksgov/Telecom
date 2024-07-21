import os
import csv
import chardet
import sys
import re
from datetime import datetime
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from design import Ui_MainWindow
from openpyxl.styles import Alignment, Font, NamedStyle, Border, Side
from PyQt5.QtGui import QFont, QIcon
from PyQt5.QtCore import Qt
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtWidgets import QVBoxLayout, QDialog, QComboBox, QHBoxLayout
import ctypes

myappid = 'Telecom'
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

def show_custom_message_box(parent, title, message, icon_path=None):
    msg_box = QMessageBox(parent)
    msg_box.setWindowTitle(title)
    msg_box.setText(message)

    if icon_path:
        msg_box.setIconPixmap(QIcon(icon_path).pixmap(64, 64))
    else:
        msg_box.setIcon(QMessageBox.Information)

    font = QFont("Century Gothic", 12)
    msg_box.setFont(font)

    msg_box.setStyleSheet("""
        QMessageBox {
            background-color: rgb(222, 241, 255);
            border-radius: 10px;
        }
        QLabel {
            color: rgb(30, 74, 163);
        }
        QPushButton {
            background-color: rgb(88, 176, 226);
            color: white;
            border-radius: 5px;
            padding: 5px 15px;
            font-weight: bold;
        }
        QPushButton:hover {
            background-color: rgb(70, 141, 181);
        }
        QPushButton:pressed {
                background-color: rgb(97,193,248);
            }
    """)

    msg_box.setStandardButtons(QMessageBox.Ok)
    ok_button = msg_box.button(QMessageBox.Ok)
    ok_button.setText("ОК")
    ok_button.setFocus()
    ok_button.setFocusPolicy(Qt.NoFocus)

    return msg_box.exec_()

class MyApp(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle("Telecom")
        self.setWindowIcon(QIcon('images\\telecom.ico'))

        # Подключаем кнопки к функциям
        self.ui.changeButton.clicked.connect(self.manage_employees)
        self.ui.fileButton1.clicked.connect(self.create_custom_excel)
        self.ui.fileButton2.clicked.connect(self.display_line_edit_text)
        self.ui.fileButton3.clicked.connect(self.load_file)
        self.ui.diagramButton1.clicked.connect(self.create_histogram)
        self.ui.lineEdit.returnPressed.connect(self.display_line_edit_text)
        self.ui.lineEdit.textChanged.connect(self.on_text_changed)
        self.ui.diagramButton2.clicked.connect(self.create_individual_chart)

        self.reports_folder = "Индивидуальные_отчеты"
        if not os.path.exists(self.reports_folder):
            os.makedirs(self.reports_folder)

        self.reports_folder = "Общие_отчеты"
        if not os.path.exists(self.reports_folder):
            os.makedirs(self.reports_folder)

        # Пути к файлам Excel
        self.excel_file = "Работники.xlsx"
        self.custom_excel_file = None
        self.report_excel_file = "ОТЧЕТ_ТЕЛЕКОМ.xlsx"
        self.report_date = None

    def on_text_changed(self):
        if self.ui.lineEdit.text():
            self.ui.lineEdit.setStyleSheet("color: black;")
        else:
            self.ui.lineEdit.setStyleSheet("color: rgb(146,146,146);")

    def create_histogram(self):
        try:
            if not os.path.exists(self.custom_excel_file):
                raise FileNotFoundError(f"Файл {self.custom_excel_file} не найден")

            wb = load_workbook(self.custom_excel_file, data_only=True)
            ws = wb["Отчет"]

            workers_wb = load_workbook('Работники.xlsx', data_only=True)
            workers_ws = workers_wb.active

            total_without_nds_col = None
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header == "Итого без НДС":
                    total_without_nds_col = col
                if total_without_nds_col:
                    break

            if not total_without_nds_col:
                raise ValueError("Не удалось найти нужные столбцы в файле")

            totals = []
            fio_dict = {}
            limit_dict = {}

            for row in workers_ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 4:
                    fio = row[0]
                    number = row[1]
                    limit = row[3]
                    if fio and number and limit:
                        fio_dict[str(number)] = fio
                        limit_dict[str(number)] = float(limit)

            for row in range(2, ws.max_row):
                total_without_nds = ws.cell(row=row, column=total_without_nds_col).value
                if total_without_nds is not None:
                    try:
                        total_value = float(total_without_nds)
                        total_with_nds = total_value * 1.2
                        totals.append(total_with_nds)
                    except ValueError:
                        print(f"Не удалось преобразовать значение '{total_without_nds}' в число")

            if not totals:
                raise ValueError("Не удалось собрать данные для построения графика")

            data = list(zip(fio_dict.values(), fio_dict.keys(), totals))

            data.sort(key=lambda x: x[0])

            sorted_fio, sorted_numbers, sorted_totals = zip(*data)

            dialog = QDialog(self)
            dialog.setWindowTitle("Гистограмма расходов")
            dialog.setGeometry(100, 100, 1800, 800)

            fig, ax = plt.subplots(figsize=(10, 6))
            bars = ax.bar(range(len(sorted_fio)), sorted_totals)
            ax.set_xlabel("Абоненты")
            ax.set_ylabel("Итого с НДС")
            ax.set_title("Расходы по абонентам")
            ax.set_xticks(range(len(sorted_fio)))
            ax.set_xticklabels(sorted_fio, rotation=90, ha='right')

            # Добавление линии лимита
            for i, (fio, number, total) in enumerate(zip(sorted_fio, sorted_numbers, sorted_totals)):
                limit = limit_dict.get(number.split()[0], 0)
                ax.plot([i - 0.4, i + 0.4], [limit, limit], color='#B0333A', linestyle='--', linewidth=2)

                if total > limit:
                    bars[i].set_color('#DC7077')
                else:
                    bars[i].set_color('#3384B0')

            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2., height,
                        f'{height:.2f}',
                        ha='center', va='bottom', rotation=0)

            ax.plot([], [], color='#B0333A', linestyle='--', linewidth=2, label='Сумма лимита с НДС')
            ax.legend(loc='lower left', bbox_to_anchor=(-0.1, -0.15))

            plt.tight_layout()

            canvas = FigureCanvas(fig)

            layout = QVBoxLayout()
            layout.addWidget(canvas)
            dialog.setLayout(layout)

            dialog.exec_()

        except Exception as e:
            error_message = f"Не удалось создать гистограмму: {str(e)}\n\n"
            error_message += "Дополнительная информация:\n"
            error_message += f"Файл: {self.custom_excel_file}\n"
            error_message += f"Тип ошибки: {type(e).__name__}"
            show_custom_message_box(self, "Ошибка", error_message)

    def manage_employees(self):
        if not os.path.exists(self.excel_file):
            self.create_new_excel()
            show_custom_message_box(self, "Информация", "Создан новый файл Работники.xlsx")

        self.open_excel_file(self.excel_file)

    def create_new_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Список работников"

        headers = ["ФИО", "Номер", "Должность", "Сумма лимита руб. с НДС", "Счет затрат"]

        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        wb.save(self.excel_file)

    def open_excel_file(self, filename):
        try:
            if os.name == 'nt':
                os.startfile(filename)
            elif os.name == 'posix':
                os.system(f"open {filename}")
        except Exception as e:
            show_custom_message_box(self, "Ошибка", f"Не удалось открыть файл: {e}")

    def load_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Выберите файл", "",
                                                   "CSV Files (*.csv);;Excel Files (*.xlsx *.xls)")
        if file_name:
            print(f"Выбран файл: {file_name}")
            self.ui.fileLabel.setText(f"Файл: {os.path.basename(file_name)}")

            try:
                if file_name.endswith('.csv'):
                    self.convert_csv_to_excel(file_name)
                elif file_name.endswith(('.xlsx', '.xls')):
                    wb = load_workbook(file_name)
                    print(f"Открыт Excel файл, листов: {len(wb.sheetnames)}")
            except Exception as e:
                print(f"Ошибка при обработке файла: {e}")
                show_custom_message_box(self, "Ошибка", f"Не удалось обработать файл: {e}")

    def convert_csv_to_excel(self, csv_file):
        excel_file = "ОТЧЕТ_ТЕЛЕКОМ.xlsx"

        wb = Workbook()
        ws = wb.active

        try:
            with open(csv_file, 'rb') as f:
                result = chardet.detect(f.read())
            encoding = result['encoding']

            with open(csv_file, newline='', encoding=encoding) as f:
                reader = csv.reader(f, delimiter=';')
                data = list(reader)

            for row in data:
                ws.append(row)

            columns_to_delete = []
            for col in range(1, ws.max_column + 1):
                if all(ws.cell(row=row, column=col).value in (0, '0', None) for row in range(2, ws.max_row + 1)):
                    columns_to_delete.append(col)

            for col in sorted(columns_to_delete, reverse=True):
                ws.delete_cols(col)

            wb.save(excel_file)
            show_custom_message_box(self, "Информация", f"CSV файл преобразован в Excel: {excel_file}")

        except Exception as e:
            print(f"Ошибка при чтении CSV файла: {e}")
            show_custom_message_box(self, "Ошибка", f"Не удалось прочитать CSV файл: {e}")

    def create_custom_excel(self):
        try:
            workers_wb = load_workbook('Работники.xlsx', data_only=True)
            workers_ws = workers_wb.active

            report_wb = load_workbook(self.report_excel_file, data_only=True)
            report_ws = report_wb.active

            start_period_col = None
            for col in range(1, report_ws.max_column + 1):
                header = report_ws.cell(row=1, column=col).value
                if header == "Начало периода":
                    start_period_col = col
                    break

            if not start_period_col:
                raise ValueError("Не удалось найти столбец 'Начало периода' в файле")

            start_period = report_ws.cell(row=2, column=start_period_col).value
            if isinstance(start_period, datetime):
                month_year = start_period.strftime("%m.%y")
            else:
                try:
                    start_date = datetime.strptime(start_period, "%d.%m.%Y")
                    month_year = start_date.strftime("%m.%y")

                except ValueError:
                    month_year = "unknown"

            self.report_date = month_year
            print(self.report_date)
            file_name = f"ОБЩИЙ_ОТЧЕТ_{month_year}.xlsx"
            output_file = os.path.join(self.reports_folder, file_name)
            self.custom_excel_file = output_file

            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет"

            headers = ["АБОНЕНТ", "Итого без НДС", "Сумма НДС", "Итого с НДС"]
            ws.append(headers)
            columns = {cell.value: cell.column for cell in report_ws[1] if
                       cell.value in ["АБОНЕНТ", "Итого без НДС", "Сумма НДС", "Итого с НДС"]}

            for row in range(2, report_ws.max_row + 1):
                data_row = []
                for header in headers:
                    col = columns.get(header)
                    if col:
                        cell_value = report_ws.cell(row=row, column=col).value
                        if header == "Итого без НДС" and cell_value is not None:
                            try:
                                cell_value = float(str(cell_value).replace(',', '.'))
                            except ValueError:
                                cell_value = 0
                        data_row.append(cell_value)
                    else:
                        data_row.append(None)

                ws.append(data_row)

            sum_nds_column = get_column_letter(headers.index("Сумма НДС") + 1)
            total_with_nds_column = get_column_letter(headers.index("Итого с НДС") + 1)
            for row_idx in range(2, ws.max_row + 1):
                sum_nds_cell = ws[f"{sum_nds_column}{row_idx}"]
                sum_without_nds_cell = ws[f"B{row_idx}"]
                total_with_nds_cell = ws[f"{total_with_nds_column}{row_idx}"]
                sum_nds_cell.value = f"={sum_without_nds_cell.coordinate} * 0.2"
                total_with_nds_cell.value = f"={sum_without_nds_cell.coordinate} + {sum_nds_cell.coordinate}"

            last_row = ws.max_row + 2
            ws.cell(row=last_row, column=1, value="Сумма:")

            for col in range(2, ws.max_column + 1):
                col_letter = get_column_letter(col)
                ws.cell(row=last_row, column=col, value=f"=SUM({col_letter}2:{col_letter}{last_row - 1})")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')

            for col_idx, header in enumerate(headers, start=1):
                column_letter = get_column_letter(col_idx)
                column_width = max(len(str(header)) + 2, 10)
                ws.column_dimensions[column_letter].width = column_width
            ws.column_dimensions['A'].width = 15

            second_ws = wb.create_sheet(title="Подробный отчет")

            start_period_col = None
            end_period_col = None
            for col in range(1, report_ws.max_column + 1):
                header = report_ws.cell(row=1, column=col).value
                if header == "Начало периода":
                    start_period_col = col
                elif header == "Конец периода":
                    end_period_col = col
                if start_period_col and end_period_col:
                    break

            start_period = report_ws.cell(row=2, column=start_period_col).value if start_period_col else "01"
            end_period = report_ws.cell(row=2, column=end_period_col).value if end_period_col else "30"

            start_date = start_period.strftime("%d.%m.%Y") if isinstance(start_period, datetime) else start_period
            end_date = end_period.strftime("%d.%m.%Y") if isinstance(end_period, datetime) else end_period

            data = [
                ["Сведения о расходовании денежных средств на мобильную связь Ухтинский филиал"],
                [f"за период с {start_date} по {end_date} г."],
                ["Номер телефона", "ФИО", "Должность", "Сумма лимита руб. с НДС",
                 "Фактическая сумма Руб. с НДС", "Фактическая сумма Руб.без НДС",
                 "Перерасход", "Счет затрат", "Тариф", "Счет 20 с НДС", "Счет 20 без НДС",
                 "Счет 26 С НДС", "Счет 26 без НДС"]
            ]

            all_border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

            for row in data:
                second_ws.append(row)

            total_with_nds_col = next(cell.column for cell in ws[1] if cell.value == "Итого с НДС")
            fact_sum_col = 5
            for row in range(2, ws.max_row - 1):
                value = ws.cell(row=row, column=total_with_nds_col).value
                if value is not None:
                    second_ws.cell(row=row + 2, column=fact_sum_col, value=value)

            for row in range(4, second_ws.max_row + 1):
                cell = second_ws.cell(row=row, column=fact_sum_col)
                cell.value = f'={ws.title}!{get_column_letter(total_with_nds_col)}{row - 2}'

            limit_col = 4
            overspend_col = 7
            for row in range(4, second_ws.max_row + 1):
                overspend_cell = second_ws.cell(row=row, column=overspend_col)
                formula = (f'=IF({get_column_letter(fact_sum_col)}{row}-{get_column_letter(limit_col)}{row}>0,'
                           f'{get_column_letter(fact_sum_col)}{row}-{get_column_letter(limit_col)}{row},"—")')
                overspend_cell.value = formula

            def apply_formula(sheet, start_row, col_condition, col_result, condition_value, col_to_copy):
                for row in range(start_row, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=col_result)
                    cell.value = f'=IF({get_column_letter(col_condition)}{row}={condition_value}, {get_column_letter(col_to_copy)}{row}, "")'

            fact_sum_nds = 5  # "Фактическая сумма Руб. с НДС"
            fact_sum_no_nds = 6  # "Фактическая сумма Руб. без НДС"
            account_col = 8  # "Счет затрат"
            nds_with_20 = 10  # "Счет 20 с НДС"
            nds_without_20 = 11  # "Счет 20 без НДС"
            nds_with_26 = 12  # "Счет 26 с НДС"
            nds_without_26 = 13  # "Счет 26 без НДС"

            apply_formula(second_ws, 4, account_col, nds_with_20, 20, fact_sum_nds)
            apply_formula(second_ws, 4, account_col, nds_without_20, 20, fact_sum_no_nds)
            apply_formula(second_ws, 4, account_col, nds_with_26, 26, fact_sum_nds)
            apply_formula(second_ws, 4, account_col, nds_without_26, 26, fact_sum_no_nds)

            second_ws.merge_cells('A1:J1')
            second_ws.merge_cells('A2:J2')

            def header_style(ws, cell):
                ws[cell].alignment = Alignment(horizontal='center')
                ws[cell].font = Font(name='Arial Cyr', bold=True, size=12)

            def title_style(ws, cell):
                ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws[cell].font = Font(name='Arial Cyr', size=9)

            def regular_style(ws, cell):
                ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws[cell].font = Font(name='Liberation Serif', size=12)

            header_style(second_ws, 'A1')
            header_style(second_ws, 'A2')

            for col in range(1, 14):
                cell = second_ws.cell(row=3, column=col)
                title_style(second_ws, cell.coordinate)

            for row in second_ws.iter_rows():
                for cell in row:
                    if cell.coordinate not in ['A1', 'A2'] and not ('A3' <= cell.coordinate <= 'N3'):
                        regular_style(second_ws, cell.coordinate)
                    if 'J3' <= cell.coordinate <= 'M3':
                        cell.alignment = Alignment(vertical='bottom', horizontal='center')

            for row in second_ws.iter_rows(min_row=4, min_col=1):
                for cell in row:
                    cell.font = Font(size=12)

            for row in range(10, 30):
                cell = second_ws.cell(row=row, column=10)
                regular_style(second_ws, cell.coordinate)

            cell = second_ws.cell(row=9, column=13)
            regular_style(second_ws, cell.coordinate)

            column_widths = [125, 135, 240, 111, 111, 111, 80, 80, 80, 101, 101, 101, 101]  # Ширина столбцов
            for i, width in enumerate(column_widths, start=1):
                excel_width = width / 7  # Преобразуем пиксели в "экселевские" единицы
                second_ws.column_dimensions[get_column_letter(i)].width = excel_width

            second_ws.row_dimensions[3].height = 38
            phone_style = NamedStyle(name="phone_style")
            phone_style.number_format = '[<=9999999]###-####;(###) #-##-##'
            phone_style.alignment = Alignment(horizontal="left")

            abonents = []
            for row in report_ws.iter_rows(min_row=2, min_col=columns["АБОНЕНТ"], max_col=columns["АБОНЕНТ"],
                                           max_row=report_ws.max_row):
                cell_value = row[0].value
                if cell_value is not None:
                    phone_number = f"{int(cell_value):010}"
                    formatted_number = f"({phone_number[:5]}) {phone_number[5]}-{phone_number[6:8]}-{phone_number[8:]}"
                    abonents.append(formatted_number)

            for idx, abonent in enumerate(abonents, start=1):
                second_ws.cell(row=idx + 3, column=1, value=abonent)

            fio_column = None
            position_column = None
            limit_column = None
            for cell in workers_ws[1]:
                if cell.value == "ФИО":
                    fio_column = cell.column
                if cell.value == "Должность":
                    position_column = cell.column
                if cell.value == "Сумма лимита руб. с НДС":
                    limit_column = cell.column
                if cell.value == "Счет затрат":
                    cost_column = cell.column

            if fio_column is not None:
                for row_idx, row in enumerate(workers_ws.iter_rows(min_row=2, min_col=fio_column, max_col=fio_column, max_row=workers_ws.max_row), start=1):
                    second_ws.cell(row=row_idx + 3, column=2, value=row[0].value)

            if position_column is not None:
                for row_idx, row in enumerate(
                        workers_ws.iter_rows(min_row=2, min_col=position_column, max_col=position_column,
                                             max_row=workers_ws.max_row), start=1):
                    cell = second_ws.cell(row=row_idx + 3, column=3, value=row[0].value)

            if limit_column is not None:
                for row_idx, row in enumerate(workers_ws.iter_rows(min_row=2, min_col=limit_column, max_col=limit_column, max_row=workers_ws.max_row), start=1):
                    second_ws.cell(row=row_idx + 3, column=4, value=row[0].value)

            if cost_column is not None:
                for row_idx, row in enumerate(workers_ws.iter_rows(min_row=2, min_col=cost_column, max_col=cost_column, max_row=workers_ws.max_row), start=1):
                    second_ws.cell(row=row_idx + 3, column=8, value=row[0].value)

            for row in report_ws.iter_rows(min_row=2, min_col=columns["Итого без НДС"],
                                           max_col=columns["Итого без НДС"], max_row=report_ws.max_row):
                cell_value = row[0].value
                if cell_value is not None:
                    try:
                        cell_value = float(str(cell_value).replace(',', '.'))
                    except ValueError:
                        cell_value = 0
                second_ws.cell(row=row[0].row + 2, column=6, value=cell_value)

            last_row = second_ws.max_row + 1
            second_ws.cell(row=last_row, column=1, value="Итого Ухтинский филиал:")
            second_ws.merge_cells(f'A{last_row}:B{last_row}')

            total_cell = second_ws.cell(row=last_row, column=1)
            total_cell.font = Font(name='Arial Cyr', bold=True, size=12)
            total_cell.alignment = Alignment(horizontal='right', vertical='center')

            for col in [4, 5, 6, 10, 11, 12, 13]:
                cell = second_ws.cell(row=last_row, column=col)
                cell.value = f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{last_row - 1})'
                cell.font = Font(name='Arial Cyr', bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '#,##0.00'

            for col in range(1, 14):
                cell = second_ws.cell(row=last_row, column=col)
                cell.border = all_border

            overspend_row = last_row + 2
            second_ws.cell(row=overspend_row, column=1, value="К удержанию из зарплаты:")
            second_ws.merge_cells(f'A{overspend_row}:B{overspend_row}')
            second_ws.merge_cells(f'C{overspend_row}:F{overspend_row}')
            second_ws.cell(row=overspend_row, column=7, value=f'=SUM(G4:G{last_row - 1})')

            overspend_cell = second_ws.cell(row=overspend_row, column=7)
            overspend_cell.font = Font(name='Arial Cyr', bold=True, size=12)
            overspend_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in range(1, 7):
                cell = second_ws.cell(row=overspend_row, column=col)
                cell.font = Font(name='Arial Cyr', bold=True, size=12)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = all_border
            second_ws.cell(row=overspend_row, column=7).number_format = '#,##0.00'

            result_row = overspend_row + 2
            second_ws.merge_cells(f'A{result_row}:C{result_row}')
            result_cell_name = second_ws.cell(row=result_row, column=1,
                                              value="Превышение/недорасход лимита по филиалу:")
            result_sum_cell = second_ws.cell(row=result_row, column=6)
            result_sum_cell.value = f'=E{last_row}-D{last_row}'
            result_cell = second_ws.cell(row=result_row, column=7)
            result_cell.value = (f'=IF(G{overspend_row}>0, "Перерасход", '
                                 f'"Недорасход")')
            result_cell_name.font = Font(name='Arial Cyr', bold=True, size=12)
            result_cell_name.alignment = Alignment(horizontal='left', vertical='center')
            result_sum_cell.font = Font(name='Arial Cyr', size=12)
            result_sum_cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in second_ws.iter_rows(min_row=3, min_col=1, max_row=result_row, max_col=13):
                for cell in row:
                    if not (cell.row == 3 and 10 <= cell.column <= 13):
                        cell.border = all_border

            for row in second_ws.iter_rows(min_row=4, max_row=last_row - 1, min_col=1, max_col=9):
                for cell in row:
                    regular_style(second_ws, cell.coordinate)

            wb.save(output_file)
            show_custom_message_box(self, "Информация", f"Создан новый файл {file_name}")
            self.open_excel_file(output_file)

        except Exception as e:
            print(f"Ошибка при создании общего отчета: {e}")
            show_custom_message_box(self, "Ошибка", f"Не удалось создать общий отчет: {e}")

    def display_line_edit_text(self):
        input_fio = self.ui.lineEdit.text().strip()

        def extract_month_year(filename):
            match = re.search(r'ОБЩИЙ_ОТЧЕТ_(\d{2})\.(\d{2})\.xlsx', filename)
            if match:
                return f"{match.group(1)}.{match.group(2)}"
            return None

        def apply_style(ws, cell, font_name='Arial Cyr', font_size=9, is_title=True):
            ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[cell].font = Font(name=font_name, size=font_size)
            if not is_title:
                ws[cell].font = Font(name='Liberation Serif', size=12)

        def set_borders(ws, start_col, end_col, start_row, end_row):
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
            for col in range(start_col, end_col + 1):
                for row in range(start_row, end_row + 1):
                    cell = f"{get_column_letter(col)}{row}"
                    ws[cell].border = border

        if not input_fio or input_fio == "Введите ФИО":
            show_custom_message_box(self, "Информация", "Пустой запрос")
            return

        reports_folder = "Общие_отчеты"
        matching_rows = []

        try:
            for filename in os.listdir(reports_folder):
                if filename.startswith("ОБЩИЙ_ОТЧЕТ_") and filename.endswith(".xlsx"):
                    file_path = os.path.join(reports_folder, filename)
                    month_year = extract_month_year(filename)

                    wb = load_workbook(file_path, data_only=True)
                    ws = wb["Подробный отчет"]

                    headers = {cell.value: idx for idx, cell in enumerate(ws[3]) if cell.value}
                    fio_col = headers.get('ФИО')
                    limit_col = headers.get('Сумма лимита руб. с НДС')
                    fact_sum_no_nds_col = headers.get('Фактическая сумма Руб.без НДС')

                    if fio_col is None or limit_col is None or fact_sum_no_nds_col is None:
                        show_custom_message_box(self, "Ошибка",
                                                f"Не найдены все необходимые столбцы в файле {filename}")
                        continue

                    for row in ws.iter_rows(min_row=4, values_only=True):
                        if row[fio_col] and row[fio_col].strip().lower() == input_fio.lower():
                            fact_sum_no_nds = row[fact_sum_no_nds_col] or 0
                            fact_sum_with_nds = fact_sum_no_nds * 1.2
                            limit = row[limit_col] or 0
                            overspend = max(fact_sum_with_nds - limit, 0)

                            new_row = list(row[:fio_col]) + list(row[fio_col + 1:fact_sum_no_nds_col - 1]) + [
                                fact_sum_with_nds, fact_sum_no_nds, overspend, month_year]
                            matching_rows.append(new_row)

            if matching_rows:
                new_wb = Workbook()
                new_ws = new_wb.active
                new_ws.row_dimensions[1].height = 25

                column_widths = [20, 30, 25, 30, 30, 20, 15]

                headers = ["Номер телефона", "Должность", "Сумма лимита руб. с НДС",
                           "Фактическая сумма Руб. с НДС", "Фактическая сумма Руб. без НДС",
                           "Перерасход", "Дата"]

                new_ws.insert_rows(1)
                new_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                new_ws.cell(row=1, column=1, value=f"{input_fio}").alignment = Alignment(horizontal='center',
                                                                                              vertical='center')
                new_ws.row_dimensions[1].height = 30
                new_ws['A1'].font = Font(name='Arial Cyr', size=14, bold=True)

                for col, header in enumerate(headers, start=1):
                    cell = new_ws.cell(row=2, column=col, value=header)
                    apply_style(new_ws, cell.coordinate, is_title=True)
                    new_ws.column_dimensions[cell.column_letter].width = column_widths[col - 1]

                for row_index, row_data in enumerate(matching_rows, start=3):
                    new_ws.row_dimensions[row_index].height = 20
                    for col, value in enumerate(row_data, start=1):
                        cell_ref = new_ws.cell(row=row_index, column=col, value=value)
                        apply_style(new_ws, cell_ref.coordinate, is_title=False)
                        if col in [3, 4, 5, 6]:
                            cell_ref.number_format = '#,##0.00'

                set_borders(new_ws, start_col=1, end_col=7, start_row=2, end_row=len(matching_rows) + 2)

                file_name = f"Индивидуальные_отчеты\\{input_fio}_отчет.xlsx"
                new_wb.save(file_name)
                show_custom_message_box(self, "Информация", f"Файл {file_name} создан успешно. Данные записаны.")
                self.open_excel_file(file_name)
            else:
                show_custom_message_box(self, "Информация", "ФИО не найдено ни в одном из отчетов")

        except Exception as e:
            show_custom_message_box(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def create_individual_chart(self):
        try:
            input_fio = self.ui.lineEdit.text().strip()

            if not input_fio or input_fio == "Введите ФИО":
                show_custom_message_box(self, "Информация", "Пожалуйста, введите ФИО")
                return

            file_path = f"Индивидуальные_отчеты\\{input_fio}_отчет.xlsx"

            if not os.path.exists(file_path):
                show_custom_message_box(self, "Ошибка", f"Отчет для {input_fio} не найден")
                return

            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            phone_numbers = set()
            for row in ws.iter_rows(min_row=3, min_col=1, max_col=1, values_only=True):
                if row[0]:
                    phone_numbers.add(row[0])

            chart_dialog = QDialog(self)
            chart_dialog.setWindowTitle(f"График расходов {input_fio}")
            chart_dialog.setGeometry(100, 100, 1500, 600)

            combo_box = QComboBox()
            combo_box.addItems(sorted(phone_numbers))

            fig, ax = plt.subplots(figsize=(10, 6))
            canvas = FigureCanvas(fig)

            annot = ax.annotate("", xy=(0, 0), xytext=(20, 20), textcoords="offset points",
                                bbox=dict(boxstyle="round", fc="w"),
                                arrowprops=dict(arrowstyle="->"))
            annot.set_visible(False)

            def update_annot(ind, x, y):
                annot.xy = (x, y)
                text = f"Дата: {x.strftime('%m.%y')}\nСумма: {y:.2f}"
                annot.set_text(text)
                annot.get_bbox_patch().set_alpha(0.4)

            def hover(event):
                vis = annot.get_visible()
                if event.inaxes == ax:
                    for line in ax.get_lines():
                        cont, ind = line.contains(event)
                        if cont:
                            x, y = line.get_data()
                            annot.xy = (x[ind["ind"][0]], y[ind["ind"][0]])
                            update_annot(ind["ind"][0], x[ind["ind"][0]], y[ind["ind"][0]])
                            annot.set_visible(True)
                            fig.canvas.draw_idle()
                            return
                if vis:
                    annot.set_visible(False)
                    fig.canvas.draw_idle()

            fig.canvas.mpl_connect("motion_notify_event", hover)

            def update_chart(selected_number):
                ax.clear()
                dates = []
                amounts = []
                limits = []
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if row[0] == selected_number:
                        date = row[-1]
                        amount = row[4]
                        limit = row[2]
                        if date and amount and limit:
                            dates.append(datetime.strptime(date, "%m.%y"))
                            amounts.append(float(amount))
                            limits.append(float(limit))

                ax.plot(dates, amounts, marker='o', label='Фактическая сумма')
                ax.plot(dates, limits, color='red', linestyle='--', label='Лимит')

                ax.set_xlabel("Дата")
                ax.set_ylabel("Сумма (руб. с НДС)")
                ax.set_title(f"Динамика расходов для {input_fio} (номер: {selected_number})")
                ax.legend()

                ax.xaxis.set_major_formatter(DateFormatter('%m.%y'))
                plt.xticks(rotation=45)
                plt.tight_layout()

                canvas.draw()

            combo_box.currentTextChanged.connect(update_chart)

            layout = QVBoxLayout()
            top_layout = QHBoxLayout()
            top_layout.addWidget(combo_box, alignment=Qt.AlignLeft)
            top_layout.addStretch(1)
            layout.addLayout(top_layout)
            layout.addWidget(canvas)

            chart_dialog.setLayout(layout)

            update_chart(combo_box.currentText())

            chart_dialog.exec_()

        except Exception as e:
            show_custom_message_box(self, "Ошибка", f"Не удалось создать график: {str(e)}")


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MyApp()
    window.show()
    sys.exit(app.exec_())